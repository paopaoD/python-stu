'''
    excel
'''
import datetime

from openpyxl import load_workbook,Workbook

###1 创建：
# 创建一个Excel
wb = Workbook()

# 获取当前active的sheet名
sheet = wb.active
print(sheet.title) # 打印Sheet表名

sheet.title = "paopao"  # 修改sheet名为 paopao


###2 打开已有文件
# wb2 = load_workbook("test_excel.xlsx")
# print(wb2.active)   # 打印打开的文件active的sheet名


###3 写数据
#方式一：   数据可以直接分配到单元格中
sheet["B3"] = "hello paopao"
sheet["C3"] = "12234548"

#方式二：   可以附加行，从第一列开始附加(从最下方空白处，第一列开始)(可以输入多行)
sheet.append(["jame",2,3])  # 相当于是增加了A1,B1,C1,3个数据

#方式三：   python类型会被转换，比如增加时间类型
sheet["A5"] = datetime.datetime.now().strftime("%Y-%m-%d")



wb.save("test_excel.xlsx")  # 保存Excel文件

























