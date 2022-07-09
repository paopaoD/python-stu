'''
    打开已有的excel
'''
from openpyxl import load_workbook

#### 1 打开已有的表格
wb = load_workbook("test_excel.xlsx")

## 打印表格内所有的sheet名
print(wb.sheetnames)    # ['paopao', 'Sheet1']



print("----------------------获取-------------------------")
#### 2 获取某一个sheet
name = wb["paopao"]
print(name) # <Worksheet "paopao">

# 打印某个单元格
print(name["C3"])   # <Cell 'paopao'.B1>

# 获取表内某个单元格的值
print(name["A1"].value)   # 12234548

# 获取指定列的切片数据
for cell in name["A1:A4"]:
    print(cell[0].value)



print("----------------------修改-------------------------")
#### 3 修改单元格内容
name["A8"] = "修改内容" # 原单元格为空，直接添加
name["F6"] = "888" # 原单元格不为空，直接覆盖

wb.save("test_excel01.xlsx")    # 保存到新excel表中



print("----------------------遍历-------------------------")
#### 4 遍历

print("---行---")
# 按 行 遍历表格，获取所有内容
for row in name:    # 循环获取表数据
    for cell in row:    # 循环获取每个单元格数据
        print(cell.value,end="\t")
    print()


print("---列---")
# 按 列 遍历表格，获取所有内容
for column in name.columns:
    print()
    for cell in column:
        print(cell.value,end=" ")
    print()


print("---指定的行&列---")
# 遍历指定的行&列
                # 从第2行开始至第5行，每行打印3列
for row in name.iter_rows(min_row=2,max_row=5,max_col=3):
    for cell in row:
        print(cell.value,end="\t")
    print()


print("---指定的列---")
# 遍历指定列的数据
        # 获取第2-5列的数据
for col in name.iter_cols(min_col=2,max_col=5,min_row=3,max_row=4):
    for item in col:
        print(item.value,end=",")
    print()


print("----------------------删除-------------------------")
#### 5 删除
print(wb.sheetnames)
# wb.remove("Sheet1")
del wb["Sheet1"]    # 删除表格
wb.save("test_excel.xlsx")  # 保存

print(wb.active)




















































