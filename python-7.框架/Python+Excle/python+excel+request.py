# @Project   : Python
# @File      : python+excel+request.py
# @Auther    : 等到秋叶金黄
# @Time      : 2022/6/23, 10:38
#
import requests
from openpyxl import load_workbook,Workbook


# # 新建表
# wb = Workbook()

# 获取Excel的路径
file_path = r"C:\Users\admin\Desktop\Python\test自动化\Python+Excle\测试用例.xlsx"

# 读取 Excel 文件
excel_file = load_workbook(file_path)

# 获取Excel文件 第一页的 所有信息
excel_cases = excel_file.worksheets[0]

# 获取Excel文件第一页的 测试用例数量
count = excel_cases.max_row - 3
print("测试用例数量：",count)


# 读取每一个用例的信息
for row in range(4,excel_cases.max_row ):
    # print(row)
    # 1
    api_url = excel_cases.cell(row,3).value     # 读取接口地址 每行第3列

    api_method = excel_cases.cell(row,4).value  # 读取请求方式 每行第4列

    api_params = eval(excel_cases.cell(row,5).value)   # 读取请求参数 每行第5列

    expected_code = int(excel_cases.cell(row,6).value)   # 读取预期状态码 每行第6列

    expected_result = excel_cases.cell(row,7).value   # 读取预期结果 每行 第7列

    print(api_url,api_method,api_params)


    #2.python发起http请求 并且拿到响应值
    response = requests.request(
        url=api_url,
        method=api_method,
        params=api_params,
    )

    # 打印响应码，响应体
    print(response.status_code," ",response.text)


    #2.1 把接口返回内容写入Excel
    excel_cases.cell(row, 8).value = response.text


    #3.对接口响应内容进行判断 【断言】【状态码】【响应内容】
    if response.status_code != expected_code or (
            expected_result is not None and expected_result != response.text
    ):
        # 如果获取的实际响应码和预期状态码不一致，或 获取的实际结果和预期结果不一样
        excel_cases.cell(row, 9).value = "执行不通过！"
    else:
        excel_cases.cell(row, 9).value = "执行通过！"


# 把内存中excel数据保存
excel_file.save(r"C:\Users\admin\Desktop\Python\test自动化\Python+Excle\result.xlsx")

excel_file.close()  # 关闭
































